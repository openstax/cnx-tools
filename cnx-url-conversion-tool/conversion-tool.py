import sys
import warnings
import requests
import json
import csv
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook

archiveUrl = "https://archive.cnx.org/"

def parseInput():
    """
    Asks the user if they want to convert module IDs or collection IDs, how they
    want to input them, and for the IDs to be converted.

    Arguments:
        None
    Returns:
        isCol - True if it is a collection ID, False for module IDs.
        moduleIdList - list of module IDs as strings
        or
        collectionId - the collection ID to be converted
        getModuleIDs - True if the user wants the module IDs

    """
    isCol = True
    prompt0 = "Do you want to input module IDs or collection IDs?"
    prompt1 = "(Type \"m\" or \"c\") "
    col = handleInputs(prompt0, ["m", "c"], prompt1)
    if (col == "m"):
        isCol = False
    elif (col == "c"):
        isCol = True

    if isCol:
        prompt3 = "Input one collection ID."
        collectionId = []
        while (len(collectionId) != 1):
            collectionId = repeatPrompt(prompt3, '', False)
        prompt4 = "Do you want all of the module IDs in this collection?"
        prompt5 = "(Type 'y' or 'n') "
        getModules = handleInputs(prompt4, ['y', 'n'], prompt5)
        if (getModules == "y"):
            getModuleIDs = True
            return isCol, collectionId, getModuleIDs
        else:
            getModuleIDs = False
            return isCol, collectionId, getModuleIDs

    else:
        prompt2 = "Do you want to import the module IDs from an excel sheet?"
        prompt5 = "(Type 'y' or 'n') "
        excel = handleInputs(prompt2, ['y', 'n'], prompt5)
        getModuleIDs = True
        if (excel == "y"):
            moduleIdList = importExcel()
        else:
            prompt3 = "Input the module IDs separated by spaces or returns.\nWhen you're done, type Ctrl-d in a new line."
            moduleIdList = []
            moduleIdList = handleInputsWithReturn(prompt3, 0, '')
        return isCol, moduleIdList, getModuleIDs

def handleInputsWithReturn(prompt0, expected, prompt1):
    """
    Allows the module IDs to be inputted to command line even when separated by
    newline characters. This allows for the user to copy a column from a spreadsheet
    and paste it into command line.
    """
    print prompt0
    lines = sys.stdin.readlines()
    moduleIds = []
    for line in lines:
        ids = line.replace(",", '').rstrip("\n").split()
        moduleIds.extend(ids)
    return moduleIds


def repeatPrompt(prompt0, prompt1 = "", notModules = True):
    """
    Repeats the prompt so that the script doesn't terminate when the user doesn't
    type an input by accident.
    """
    response = ""
    while (len(response) == 0):
        print prompt0
        if (notModules):
            response = raw_input(prompt1)
        else:
            response = raw_input(prompt1).replace(",", "").split()
    return response

def handleInputs(prompt0, expected, prompt1 = '', notModules = True):
    """
    Handles different situations for getting user input.
    """
    if not notModules: # to get module and col ids
        return repeatPrompt(prompt0, prompt1, notModules)
    else:
        if (type(expected) is int): # filenames
            inputs = []
            while (len(inputs) != expected):
                inputs = repeatPrompt(prompt0, prompt1, notModules).replace(', ', ',').split(',')
            return inputs
        else: # questions like yes or no
            response = ''
            while (response not in expected):
                response = repeatPrompt(prompt0, prompt1, notModules).lower()[0]
            return response


def importExcel():
    """
    Imports the data from an excel sheet.

    Arguments:
        None
    Returns:
        idList - a list of IDs as strings
    """

    prompt7 = "Type the file name, worksheet name, and title of the column with the IDs separated by commas."
    inputs = handleInputs(prompt7, 3, '', True)
    filename = inputs[0]
    sheetname = inputs[1]
    title = inputs[2]
    if (filename[-5:] != '.xlsx'):
        filename += '.xlsx'
    try:
        wb2 = load_workbook(filename)
    except Exception as e:
        print "The input file does not exist. Please make sure that the filename is correct."
        sys.exit(1)

    sheetNames = wb2.get_sheet_names()
    if sheetname not in sheetNames:
        output = "A worksheet with the name '%s' does not exist in the file." %sheetname
        sys.exit(output)
    else:
        ws = wb2.get_sheet_by_name(sheetname)
        column = 0
        for col in ws.columns:
            if (title == col[0].value):
                column = col
                break
        if (column == 0):
            sys.exit("A column with the title '%s' doesn't exist.") %title
        ids = []
        for row in column[1:]:
            moduleId = row.value.encode('ascii','ignore').replace(' ', '')
            ids.append(moduleId)
        return ids



def convertModuleIdList(moduleIdList):
    """
    Given a list of module IDs, converts them to the other formats by seeing
    where they redirect to on archive.cnx.org, and getting the IDs from the
    json.

    Arguments:
        moduleIdList - list of ids as strings

    Returns:
        convertedList - list of dictionaries, where each one represents an ID
            and contains all three formats.
        errorIdList - list of IDs that returned an error
        count - number of ids in the converted list
    """
    convertedList = []
    errorIdList = []
    count = 0
    if (len(moduleIdList) == 0):
        print "There were no IDs entered."
        sys.exit()
    print "Retrieving IDs..."
    for moduleId in moduleIdList:
        isLegacy = True
        if (len(moduleId) != 6):
            isLegacy = False
        if isLegacy:
            url = archiveUrl + 'content/%s/latest' % moduleId
        else:
            url = archiveUrl + 'contents/%s' % moduleId
        r = requests.get(url, allow_redirects=False)
        if (r.status_code != requests.codes.ok and r.status_code != 302):
            errorIdList.append(moduleId)
            continue
        r3 = requests.get(r.headers['location']+'.json')
        if (r3.status_code != requests.codes.ok):
            errorIdList.append(moduleId)
            continue
        responseJson = r3.json()
        moduleDict = {}
        moduleDict['Short ID'] = responseJson['shortId']
        moduleDict['Long ID'] = responseJson['id']
        moduleDict['Legacy ID'] = responseJson['legacy_id']
        convertedList.append(moduleDict)
        count += 1
    return convertedList, errorIdList, count

def convertColId(colId, getModules):
    """
    Given a collection ID, converts it to the other formats by seeing
    where it redirects to on archive.cnx.org. If getModules is true, then it
    parses the json to get all of the module IDs in the collection and calls
    convertModuleIdList().

    Arguments:
        colId - the collection ID
        getModules - true if the user wants all of the module IDs, false otherwise.

    Returns:
        convertedList - list of dictionaries, where each one represents an ID
            and contains all three formats
        errorIdList - list of IDs that returned an error
        count - number of ids in the converted list
    """

    errorIdList = []
    isLegacy = True
    if (colId[0:3].lower() != "col"):
        isLegacy = False

    if isLegacy: #legacyId
        url = 'https://archive.cnx.org/content/%s/latest' % colId
    else:
        url = 'https://archive.cnx.org/contents/%s' % colId
    r = requests.get(url, allow_redirects=False)
    if (r.status_code != requests.codes.ok and r.status_code != 302):
        errorIdList.append(colId)
        sys.exit("The Collection ID returned an error.")
    r3 = requests.get(r.headers['location']+'.json')
    if (r3.status_code != requests.codes.ok):
        errorIdList.append(colId)
        sys.exit("The Collection ID returned an error.")
    responseJson = r3.json()
    moduleDict = {}
    moduleDict['Short ID'] = responseJson['shortId']
    moduleDict['Long ID'] = responseJson['id']
    moduleDict['Legacy ID'] = responseJson['legacy_id']
    convertedList = [moduleDict]
    if getModules:
        contents = responseJson['tree']['contents']
        moduleIdList = []
        for section in contents:
            findModuleIds(section, moduleIdList)
        convertedModules, moduleErrors, count = convertModuleIdList(moduleIdList)
        errorIdList.extend(moduleErrors)
        convertedList.append(convertedModules)
        return convertedList, errorIdList, count
    return convertedList, errorIdList, 0

def findModuleIds(section, moduleIdList):
    """
    Finds the module IDs from the collection json recursively.

    Arguments:
        section - a subcollection
        moduleIdList - a list for the module IDs that this function appends to
    Returns:
        None
    """
    moduleShort = section['shortId']
    if (moduleShort != 'subcol'):
        moduleIdList.append(moduleShort)
    else:
        for subSection in section['contents']:
            findModuleIds(subSection, moduleIdList)


def printResults(idList, errorList, printModules, isCol):
    """
    Prints the results in command line.
    """
    if isCol:
        colDict = idList[0]
        print "Collection:"
        print  "Legacy ID   Long ID                                 Short ID"
        print colDict['Legacy ID'], "  ", colDict['Long ID'], "  ", colDict['Short ID']
        if printModules:
            moduleList = idList[1]
    else:
        moduleList = idList
    if (printModules):
        if (len(moduleList) != 0):
            print "Modules:"
            print "Legacy ID   Long ID                                 Short ID"
            for modId in moduleList:
                print modId['Legacy ID'], "    ", modId['Long ID'], "  ", modId['Short ID']
        printErrors(errorList)

def printErrors(errorList):
    if (len(errorList) != 0):
        print "The following IDs returned an error:"
        for errorId in errorList:
            print errorId


def parseExportInput():
    """
    Asks the user what format they want for the output.
    """
    prompt4 = "Pick the output format: csv, excel, or terminal."
    prompt5 = "(type 'c', 'e' or 't') "
    export = ''
    export = handleInputs(prompt4, ['c', 'e', 't'], prompt5, True)
    if (export == 't'):
        return 't', ''
    else:
        filename = ''
        prompt6 = 'Type the filename you want for the exported file.'
        filename = handleInputs(prompt6, 1, '', True)[0]
        if (export == 'e'):
            if (filename[-5:] != '.xlsx'):
                filename += '.xlsx'
        else:
            if (filename[-4:] != '.csv'):
                filename += '.csv'
        return export, filename

def exportAsCsv(filename, idList, errorIdList, isCol):
    """
    Exports the results as a csv file.

    Arguments:
        filename - the name for the file to be exported as
        idList - list of dictionaries, each one representing one module/collection
        errorIdList - list of IDs that returned an error
    Returns:
        None
    """
    with open(filename, 'wb') as f:
        w = csv.DictWriter(f, ['Legacy ID', 'Long ID', 'Short ID'])
        w.writeheader()
        if isCol:
            colDict = idList[0]
            w.writerow(colDict)
            moduleDictList = idList[1]
        else:
            moduleDictList = idList
        if (len(moduleDictList) == 0):
            print 'There are no converted module IDs to output.'
        else:
            w.writerows(moduleDictList)
    print "Finished writing to file!"
    f.close()
    printErrors(errorIdList)

def exportAsExcel(filename, idList, errorIdList, count, isCol):
    """
    Exports the results as an excel (.xlsx) file.

    Arguments:
        filename - the name for the file to be exported as
        idList - list of dictionaries, each one representing one module/collection
        errorIdList - list of IDs that returned an error
        count - number of module IDs
    Returns:
        None
    """

    wb = Workbook()
    dest_filename = filename
    ws1 = wb.active
    ws1.title = "Converted IDs"
    header = ['Legacy ID', 'Long ID', 'Short ID']
    for col in range(1, 4):
        _ = ws1.cell(column=col, row=1, value="%s" % header[col - 1])
    if isCol:
        colDict = idList[0]
        for col in range(1, 4):
            _ = ws1.cell(column=col, row=2, value="%s" % colDict[header[col - 1]])
        moduleDictList = idList[1]
        startRow = 3
    else:
        moduleDictList = idList
        startRow = 2
    if (len(moduleDictList) == 0):
        print 'There are no converted module IDs to output.'
    else:
        for row in range(startRow, count + startRow):
            for col in range(1, 4):
                _ = ws1.cell(column=col, row=row, value="%s" % moduleDictList[row - startRow][header[col - 1]])
    wb.save(filename = dest_filename)
    print "Finished writing to file!"
    printErrors(errorIdList)

def main():
    warnings.simplefilter('ignore', UserWarning)
    isCol, idList, printModules = parseInput()
    if (len(idList) == 0):
        sys.exit()
    export, filename = parseExportInput()
    if isCol:
        convertedList, errorIdList, count = convertColId(idList[0], printModules)
    else:
        convertedList, errorIdList, count = convertModuleIdList(idList)
    if (export == 'c'):
        exportAsCsv(filename, convertedList, errorIdList, isCol)
    elif (export == 'e'):
        exportAsExcel(filename, convertedList, errorIdList, count, isCol)
    else:
        printResults(convertedList, errorIdList, printModules, isCol)


if __name__ == '__main__':
    main()
